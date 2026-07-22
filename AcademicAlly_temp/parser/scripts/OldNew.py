import json
from openpyxl import load_workbook

FinalMap = {}

def OldAndNew():
    path = "../utilities/TIMETABLEJULYTODEC25.xlsx"
    workbook = load_workbook(filename=path, data_only=True)

    def clean_value(value):
        if value is not None:
            return str(value).strip().replace(" ", "")
        return

    for sheetName in workbook.sheetnames:
        sheet = workbook[sheetName]

        tutorialNames = {col: clean_value(sheet.cell(row=5, column=col).value) for col in range(1, sheet.max_column + 1)}
        practicalNames = {col: clean_value(sheet.cell(row=6, column=col).value) for col in range(1, sheet.max_column + 1)}

        for col in range(1, sheet.max_column + 1):
            key = tutorialNames[col]
            value = tutorialNames[col]
            #value = f"{tutorialNames[col]} / {practicalNames[col]}"
            FinalMap[key] = value

    with open('subgroup.json', 'w') as final:
        json.dump(FinalMap, final, indent=4)

OldAndNew()
