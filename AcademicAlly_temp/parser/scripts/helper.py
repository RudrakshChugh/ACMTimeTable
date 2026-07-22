import openpyxl

from openpyxl.utils import get_column_letter
def count_merged_cells_along_row(ws, row, column):


    cell_address = f"{get_column_letter(column)}{row}"


    cell = ws[cell_address]


    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:

            min_col, min_row, max_col, max_row = merged_range.bounds
            print(min_col, min_row, max_col, max_row)

            number_of_merged_cells_along_row = max_col - min_col + 1
            return number_of_merged_cells_along_row




file_path = '../utilities/TimeTable.xlsx'
sheet_name = '3RD YEAR B'


# Load the workbook and select the specified worksheet
wb = openpyxl.load_workbook(file_path)
ws = wb[sheet_name]
x=30
y=5
merged_cells_count = count_merged_cells_along_row(ws, x, y)
print(merged_cells_count)

# for m in range(y, y + merged_cells_count, 2):
#     print(ws.cell(row=x, column=m))




