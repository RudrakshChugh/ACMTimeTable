from openpyxl.utils import get_column_letter

def get_merged_cells(sheet):
    merged_cells_map = {}
    min_rows_set = set()
    merged_cells = sheet.merged_cells.ranges
    for merged_range in merged_cells:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_cell_address = f"{get_column_letter(min_col)}{min_row}"
        merged_cells_map[top_left_cell_address] = (min_col, max_col)
        min_rows_set.add(min_row)
    return merged_cells_map, min_rows_set


def remove_whitespace(s):
    if s is None:
        return None
    return ''.join(str(s).split())


def return_day(row_val):
    if 9 <= row_val <= 36:
        return "Monday"
    elif 37 <= row_val <= 64:
        return "Tuesday"
    elif 65 <= row_val <= 92:
        return "Wednesday"
    elif 93 <= row_val <= 120:
        return "Thursday"
    elif 121 <= row_val <= 148:
        return "Friday"
    elif 149 <= row_val <= 176:
        return "Saturday"
    else:
        return None


def getTypeOfClass(s):
    for char in reversed(s):
        if char != ' ':
            return char
    return None


def get_formatted_time(sheet, row, column):
    cell_value = sheet.cell(row=row, column=column).value

    if not isinstance(cell_value, str):
        try:
            cell_value = cell_value.strftime("%I:%M %p")
        except AttributeError:
            cell_value = str(cell_value)

    return cell_value.strip()



def HandlePractical(sheet, x, y, subgroup, day, time, cell, result, room, time_col):
    # room = sheet.cell(row=x + 1, column=y).value + ' ' + sheet.cell(row=x + 2, column=y).value
    result[subgroup][day][time] = [cell, room]
    print(f"result[{subgroup}][{day}][{time}] = {cell}")

    # time2 = sheet.cell(row=x + 2, column=4).value.strftime("%I:%M %p")
    time2 = get_formatted_time(sheet, x + 2, time_col)

    result[subgroup][day][time2] = [cell, room]
    print(f"result[{subgroup}][{day}][{time2}] = {cell}")
    skip_next = True


def HandleTutorial(sheet, x, y, subgroup, day, time, cell, result, room, time_col):
    tVal = None
    checker = sheet.cell(row=x + 2, column=y).value
    if (checker is not None and len(checker) <= 5):
        time2 = get_formatted_time(sheet, x + 2, time_col)
        result[subgroup][day][time2] = [cell, room]
        print(f"result[{subgroup}][{day}][{time2}] = {cell}")
        tVal = 1

    # room = sheet.cell(row=x + 1, column=y).value
    result[subgroup][day][time] = [cell, room]
    print(f"result[{subgroup}][{day}][{time}] = {cell}")
    return tVal

def count_merged_cells_along_row(merged_cells_map, cell_address):
    if cell_address in merged_cells_map:
        min_row, max_row = merged_cells_map[cell_address]
        number_of_merged_cells_along_row = max_row - min_row + 1
        return number_of_merged_cells_along_row
    return 1


def HandleLecture(sheet, x, y, subgroup, day, time, cell, result, merged_cells_map, room, time_col):
    # room = sheet.cell(row=x + 1, column=y).value
    lectureGroup = sheet.cell(row=4, column=y).value
    print(f"lectureGroup: {lectureGroup}")
    result[subgroup][day][time] = [cell, room]
    print(f"result[{subgroup}][{day}][{time}] = {cell}")
    tVal = None
    checker = sheet.cell(row=x + 2, column=y).value
    if checker is not None and len(checker) <= 5:
        time2 = get_formatted_time(sheet, x + 2, time_col)
        result[subgroup][day][time2] = [cell, room]
        print(f"result[{subgroup}][{day}][{time2}] = {cell}")
        tVal = 1
    return tVal

    # try:
    #     rangeVal = merged_cells_map[lectureGroup]
    # except KeyError:
    #     result[subgroup][day][time] = [cell, room]
    #     print(f"result[{subgroup}][{day}][{time}] = {cell}")
    #     return
    #
    # print(f"rangeVal: {rangeVal}")
    #
    # for m in range(y, rangeVal[1] + 1, 2):
    #     subgroup = remove_whitespace(sheet.cell(row=7, column=m).value)
    #     print(f"Processing subgroup at column={m}: {subgroup}")
    #     result[subgroup][day][time] = [cell, room]
    #     print(f"result[{subgroup}][{day}][{time}] = {cell}")


def parser(sheet, result):
    merged_cells_map, _ = get_merged_cells(sheet)

    row_size = sheet.max_row
    col_size = sheet.max_column

    skip_next = False
    
    start_col = 5
    time_col = 4
    start_row = 9
    
    val_at_6_5 = remove_whitespace(sheet.cell(row=6, column=5).value)
    if val_at_6_5 == "HOURS":
        start_col = 6
        time_col = 5
        
    start_row = 9
    if sheet.cell(row=9, column=time_col).value is None and sheet.cell(row=10, column=time_col).value is not None:
        start_row = 10
        
    for y in range(start_col, col_size, 2):
        for x in range(start_row, row_size, 2):

            if skip_next:
                skip_next = False
                continue

            subgroup = remove_whitespace(sheet.cell(row=6, column=y).value)

            day = return_day(x)
            if (day is None):
                break
            time = get_formatted_time(sheet, x, time_col)
            cell_address = f"{get_column_letter(y)}{x}"
            print(cell_address)
            print(sheet.title)

            cell = remove_whitespace(sheet.cell(row=x, column=y).value)
            print(sheet.cell(row=x, column=y).value)
            # Debug prints
            print(
                f"Processing cell at row={x}, column={y} | subgroup: {subgroup} | time: {time} | day: {day} |  cell value: {cell}")


            if cell is not None:
                type = getTypeOfClass(cell)
                print(f"type of class: {type}")
                if type == 'L' or type =='T':
                    room = room = sheet.cell(row=x + 1, column=y).value
                elif type == 'P':
                    cellBelow = sheet.cell(row=x + 1, column=y)
                    cellTwoBelow = sheet.cell(row=x + 2, column=y)

                    valueBelow = cellBelow.value or ""
                    valueTwoBelow = cellTwoBelow.value or ""

                    room = f"{valueBelow} {valueTwoBelow}".strip()
                    print(room)


                merged_cells_count = count_merged_cells_along_row(merged_cells_map, cell_address)
                for m in range(y, y + merged_cells_count, 2):
                    subgroup = remove_whitespace(sheet.cell(row=6, column=m).value)

                    if type == 'L':
                        tVal = HandleLecture(sheet, x, y, subgroup, day, time, cell, result, merged_cells_map, room, time_col)
                        if tVal is not None:
                            skip_next = True

                    elif type == 'P':
                        HandlePractical(sheet, x, y, subgroup, day, time, cell, result, room, time_col)
                        skip_next = True

                    elif type == 'T':
                        tVal = HandleTutorial(sheet, x, y, subgroup, day, time, cell, result, room, time_col)
                        if tVal is not None:
                            skip_next = True

